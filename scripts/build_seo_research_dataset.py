#!/usr/bin/env python3
"""Consolidate SEO research sources into a single per-page dataset.

Sources:
- database/data/seo-research/ads-search-terms.csv          (Google Ads search terms report)
- database/data/seo-research/ads-search-terms-clustered.csv (pre-clustered Ads terms)
- database/data/search-console/queries.csv                  (Google Search Console queries)
- database/data/seo-research/paa/*.json                     (People Also Ask via Apify)
- Deluxe_Windows_SEO_content_map.xlsx                       (query->page mapping, FAQ bank)

Output: database/data/seo-research/dataset.json with, for every page in
database/data/page-metadata, a ranked list of real queries plus the PAA
questions and supporting facts relevant to that page.
"""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RESEARCH = ROOT / "database" / "data" / "seo-research"
METADATA_ROOT = ROOT / "database" / "data" / "page-metadata"
GSC_CSV = ROOT / "database" / "data" / "search-console" / "queries.csv"
WORKBOOK = ROOT / "Deluxe_Windows_SEO_content_map.xlsx"

CITY_SLUGS = {
    "alameda", "antioch", "berkeley", "burlingame", "concord", "daly-city",
    "fairfield", "fremont", "hayward", "livermore", "milpitas",
    "mountain-view", "napa", "oakland", "petaluma", "piedmont", "pleasanton",
    "redwood-city", "richmond", "san-francisco", "san-jose", "san-leandro",
    "san-mateo", "san-rafael", "san-ramon", "santa-clara", "santa-rosa",
    "south-san-francisco", "sunnyvale", "vacaville", "vallejo", "walnut-creek",
}

BRAND_TOKENS = {
    "andersen": "andersen",
    "anderson": "andersen",
    "marvin": "marvin",
    "milgard": "milgard",
    "simonton": "simonton",
    "jeld wen": "jeld-wen",
    "jeld-wen": "jeld-wen",
    "jeldwen": "jeld-wen",
    "anlin": "anlin",
    "ply gem": "ply-gem",
    "plygem": "ply-gem",
    "alside": "alside",
    "western window": "western-window-systems",
    "ital windows": "italwindows",
    "italwindows": "italwindows",
    "all weather": "all-weather-architectural-aluminum",
}

MATERIAL_TOKENS = [
    ("aluminum clad", "aluminum-clad"),
    ("aluminium clad", "aluminum-clad"),
    ("wood clad", "wood-clad"),
    ("fiberglass", "fiberglass"),
    ("fibreglass", "fiberglass"),
    ("aluminum", "aluminum"),
    ("aluminium", "aluminum"),
    ("steel", "steel"),
    ("vinyl", "vinyl"),
    ("wood", "wood"),
]


def norm(query: str) -> str:
    query = query.casefold().strip()
    query = re.sub(r"[^a-z0-9\s'-]", " ", query)
    return re.sub(r"\s+", " ", query).strip()


def parse_float(value: str) -> float:
    value = value.replace(",", "").replace("%", "").replace("$", "").strip()
    try:
        return float(value)
    except ValueError:
        return 0.0


def load_ads_terms() -> dict[str, dict]:
    terms: dict[str, dict] = {}
    with open(RESEARCH / "ads-search-terms.csv", encoding="utf-8-sig") as handle:
        lines = handle.read().splitlines()
    reader = csv.DictReader(lines[2:])
    for row in reader:
        query = norm(row["Search term"])
        if not query:
            continue
        entry = terms.setdefault(
            query,
            {"clicks": 0.0, "impressions": 0.0, "cost": 0.0, "conversions": 0.0},
        )
        entry["clicks"] += parse_float(row["Clicks"])
        entry["impressions"] += parse_float(row["Impr."])
        entry["cost"] += parse_float(row["Cost"])
        entry["conversions"] += parse_float(row["Conversions"])
    return terms


def load_ads_clusters() -> dict[str, dict]:
    clusters: dict[str, dict] = {}
    with open(RESEARCH / "ads-search-terms-clustered.csv", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            query = norm(row["search_term"])
            if query and query not in clusters:
                clusters[query] = {
                    "ad_group": row["ad_group"].strip(),
                    "layer": row["layer"].strip(),
                }
    return clusters


def load_gsc() -> dict[str, dict]:
    queries: dict[str, dict] = {}
    with open(GSC_CSV, encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            query = norm(row["Top queries"])
            if not query:
                continue
            queries[query] = {
                "clicks": parse_float(row["Clicks"]),
                "impressions": parse_float(row["Impressions"]),
                "position": parse_float(row["Position"]),
            }
    return queries


def load_paa() -> dict[str, dict]:
    merged: dict[str, dict] = {}
    for file in sorted((RESEARCH / "paa").glob("*.json")):
        payload = json.loads(file.read_text(encoding="utf-8"))
        for query, data in payload["queries"].items():
            merged[norm(query)] = {
                "paa": list(data.get("paa") or []),
                "facts": list(data.get("facts") or []),
                "file": file.name,
            }
    return merged


def load_workbook_mapping() -> tuple[dict[str, list[str]], list[dict]]:
    """Return (path -> ads queries from sheet 01, FAQ bank rows from sheet 03)."""
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)

    sheet01 = list(workbook["01 SEO-карта страниц"].values)
    headers = [str(value) for value in sheet01[0]]
    query_column = headers.index("Реальные запросы из Ads (топ по показам)")
    url_column = headers.index("URL")
    page_queries: dict[str, list[str]] = {}
    for row in sheet01[1:]:
        if row[url_column] is None:
            continue
        path = "/" + str(row[url_column]).strip().strip("/")
        raw = str(row[query_column] or "")
        queries = [norm(item) for item in re.split(r"[;\n]", raw) if norm(item)]
        page_queries[path if path != "/" else "/"] = queries

    sheet03 = list(workbook["03 FAQ-банк"].values)
    bank_headers = [str(value) for value in sheet03[0]]
    faq_bank = [dict(zip(bank_headers, row)) for row in sheet03[1:] if row[1]]
    return page_queries, faq_bank


def load_pages() -> dict[str, dict]:
    pages: dict[str, dict] = {}
    for file in METADATA_ROOT.rglob("*.json"):
        record = json.loads(file.read_text(encoding="utf-8"))
        pages[str(record["path"])] = {
            "key": str(record["key"]),
            "family": str(record["key"]).split("/", 1)[0],
        }
    return pages


def detect_brand(query: str) -> str | None:
    for token, slug in BRAND_TOKENS.items():
        if token in query:
            return slug
    return None


def detect_material(query: str) -> str | None:
    for token, slug in MATERIAL_TOKENS:
        if token in query:
            return slug
    return None


def detect_city(query: str) -> str | None:
    condensed = query.replace(" ", "-")
    for city in sorted(CITY_SLUGS, key=len, reverse=True):
        if city in condensed:
            return city
    return None


def question_fallback_targets(question: str, pages: dict[str, dict]) -> list[str]:
    """Keyword-based candidate pages for a single PAA question, used when the
    seed's own targets have no free FAQ slots left."""
    query = norm(question)
    is_door = bool(re.search(r"\bdoors?\b", query))
    brand = detect_brand(query)
    material = detect_material(query)
    is_cost = bool(re.search(r"\b(cost|price|cheap|expensive|much)\b", query))

    candidates: list[str] = []
    if brand and material:
        for candidate in (
            f"/window-type/{brand}-{material}-windows",
            f"/door-types/{brand}-{material}-doors",
        ):
            candidates.append(candidate)
    if brand:
        candidates.append(f"/door-brands/{brand}" if is_door else f"/brands/{brand}")
        candidates.append(f"/brands/{brand}" if is_door else f"/door-brands/{brand}")
    if material:
        if is_door:
            candidates.append(f"/doors/{material}-doors")
        else:
            candidates.append(f"/windows/{material}-windows")
    if is_door:
        candidates.append("/doors")
    if is_cost:
        candidates.extend(["/faq", "/financing", "/special-offers"])
    candidates.extend(["/faq", "/glossary"])
    return [c for c in dict.fromkeys(candidates) if c in pages]


def rule_based_page(query: str, pages: dict[str, dict]) -> str | None:
    """Map a query to the most specific matching page path."""
    is_door = bool(re.search(r"\bdoors?\b", query))
    brand = detect_brand(query)
    material = detect_material(query)
    city = detect_city(query)

    if city and f"/window-replacement/{city}" in pages and not is_door:
        return f"/window-replacement/{city}"

    if brand and material and not is_door:
        for candidate in (
            f"/window-type/{brand}-{material}-windows",
            f"/window-type/{material}-{brand}-windows",
            f"/window-type/{brand}-{material.replace('aluminum', 'aluminium')}-windows",
        ):
            if candidate in pages:
                return candidate

    if brand:
        target = f"/door-brands/{brand}" if is_door else f"/brands/{brand}"
        if target in pages:
            return target

    if material:
        target = f"/doors/{material}-doors" if is_door else f"/windows/{material}-windows"
        if target in pages:
            return target

    return None


def main() -> None:
    ads_terms = load_ads_terms()
    ads_clusters = load_ads_clusters()
    gsc = load_gsc()
    paa = load_paa()
    page_queries, faq_bank = load_workbook_mapping()
    pages = load_pages()

    all_queries: dict[str, dict] = {}
    for query, metrics in ads_terms.items():
        all_queries[query] = {
            "query": query,
            "ads": metrics,
            "gsc": None,
            "sources": ["ads"],
        }
    for query, metrics in gsc.items():
        entry = all_queries.setdefault(
            query, {"query": query, "ads": None, "gsc": None, "sources": []}
        )
        entry["gsc"] = metrics
        entry["sources"].append("gsc")

    # Explicit workbook mapping takes priority over rules.
    explicit: dict[str, str] = {}
    for path, queries in page_queries.items():
        for query in queries:
            explicit.setdefault(query, path)

    per_page: dict[str, dict] = {
        path: {"queries": [], "paa": [], "facts": [], "paa_sources": []}
        for path in pages
    }
    unmapped: list[dict] = []

    for query, entry in all_queries.items():
        path = explicit.get(query) or rule_based_page(query, pages)
        cluster = ads_clusters.get(query)
        record = {
            "query": query,
            "ads": entry["ads"],
            "gsc": entry["gsc"],
            "layer": cluster["layer"] if cluster else None,
            "ad_group": cluster["ad_group"] if cluster else None,
            "sources": entry["sources"],
        }
        if path and path in per_page:
            per_page[path]["queries"].append(record)
        else:
            unmapped.append(record)

    def weight(record: dict) -> float:
        ads = record["ads"] or {}
        gsc_metrics = record["gsc"] or {}
        return (
            ads.get("conversions", 0.0) * 500
            + ads.get("clicks", 0.0) * 10
            + ads.get("impressions", 0.0) * 0.5
            + gsc_metrics.get("clicks", 0.0) * 20
            + gsc_metrics.get("impressions", 0.0) * 0.2
        )

    for payload in per_page.values():
        payload["queries"].sort(key=weight, reverse=True)
        payload["queries"] = payload["queries"][:40]

    # Attach PAA seeds to pages with the same routing rules. Every seed fans
    # out to its primary page plus any secondary targets so a question that is
    # already claimed by one page can still be used by the next candidate.
    paa_pool: list[dict] = []
    for seed, data in paa.items():
        primary = explicit.get(seed) or rule_based_page(seed, pages)
        targets: list[str] = []
        if primary:
            targets.append(primary)
        for extra in PAA_SEED_TARGETS.get(seed, []):
            if extra not in targets:
                targets.append(extra)
        for target in targets:
            if target in per_page:
                per_page[target]["paa"].extend(data["paa"])
                per_page[target]["facts"].extend(data["facts"])
                per_page[target]["paa_sources"].append(seed)
        for question in data["paa"]:
            question_targets = list(targets)
            for extra in question_fallback_targets(str(question), pages):
                if extra not in question_targets:
                    question_targets.append(extra)
            paa_pool.append({
                "question": question,
                "seed": seed,
                "targets": [t for t in question_targets if t in per_page],
            })

    dataset = {
        "generated_from": [
            "seo-research/ads-search-terms.csv",
            "seo-research/ads-search-terms-clustered.csv",
            "search-console/queries.csv",
            "seo-research/paa/*.json",
            "Deluxe_Windows_SEO_content_map.xlsx",
        ],
        "faq_bank": [
            {
                "cluster": str(row.get("Кластер") or ""),
                "question": str(row.get("Вопрос (дословно на страницу и в FAQPage schema)") or ""),
                "target_pages": str(row.get("Целевые страницы") or ""),
                "source_queries": str(row.get("Источник: реальные запросы (показы)") or ""),
                "answer_format": str(row.get("Формат ответа (GEO)") or ""),
            }
            for row in faq_bank
        ],
        "pages": per_page,
        "paa_pool": paa_pool,
        "unmapped_top": sorted(unmapped, key=weight, reverse=True)[:150],
        "stats": {
            "ads_queries": len(ads_terms),
            "gsc_queries": len(gsc),
            "paa_seeds": len(paa),
            "paa_pool": len(paa_pool),
            "mapped_queries": sum(len(p["queries"]) for p in per_page.values()),
            "pages_with_queries": sum(1 for p in per_page.values() if p["queries"]),
            "pages_with_paa": sum(1 for p in per_page.values() if p["paa"]),
            "unmapped_queries": len(unmapped),
        },
    }

    output = RESEARCH / "dataset.json"
    output.write_text(
        json.dumps(dataset, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(json.dumps(dataset["stats"], indent=2))


# Seeds that describe a topic rather than one page: fan them out to every
# page family that benefits from the questions. Order matters — the first
# page with a free FAQ slot claims the question.
PAA_SEED_TARGETS: dict[str, list[str]] = {
    "how much does window replacement cost": ["/faq", "/", "/windows", "/special-offers"],
    "cost to replace all windows in house": ["/faq", "/windows", "/financing"],
    "window companies bay area": ["/", "/contacts", "/testimonials", "/about"],
    "window replacement contractor": ["/contacts", "/about", "/testimonials"],
    "best window brands": ["/brand", "/windows", "/faq"],
    "how long do windows last": ["/blog/how-long-do-windows-last", "/faq", "/windows"],
    "window installation": ["/windows", "/faq", "/gallery"],
    "energy efficient windows": [
        "/blog/do-energy-efficient-windows-and-doors-make-a-difference-for-bay-area-homeowners",
        "/glossary",
        "/windows/vinyl-windows",
    ],
    "replacement windows vs new construction": [
        "/blog/new-construction-windows-vs-replacement-windows",
        "/faq",
    ],
    "do new windows increase home value": [
        "/blog/do-new-windows-increase-home-value-for-bay-area-homeowners",
        "/faq",
    ],
    "how to measure windows for replacement": [
        "/blog/how-to-measure-windows-for-replacement",
        "/faq",
    ],
    "window financing": ["/financing", "/faq", "/special-offers"],
    "dual pane windows": ["/glossary", "/windows/vinyl-windows", "/faq"],
    "window frame materials": [
        "/blog/what-kind-of-window-frame-is-right-for-you",
        "/windows",
        "/glossary",
    ],
    "entry door replacement cost": ["/doors", "/doors/fiberglass-doors", "/faq"],
    "sliding patio doors": ["/doors", "/doors/vinyl-doors", "/doors/aluminum-doors"],
    "sliding glass door replacement": ["/doors/vinyl-doors", "/doors", "/doors/aluminum-doors"],
    "fiberglass entry doors": ["/doors/fiberglass-doors", "/doors"],
    "steel entry doors": ["/doors/steel-doors", "/doors"],
    "wood entry doors": ["/doors/wood-doors", "/doors"],
    "vinyl patio doors": ["/doors/vinyl-doors", "/doors"],
    "wood clad doors": ["/doors/wood-clad-doors", "/doors"],
    "milgard tuscany series": [
        "/brand-collections/brand-milgard-v400-tuscany-series",
        "/window-type/milgard-vinyl-windows",
        "/brands/milgard",
    ],
    "milgard trinsic series": [
        "/brand-collections/brand-milgard-v300-trinsic-series",
        "/window-type/milgard-vinyl-windows",
        "/brands/milgard",
    ],
    "milgard style line series": [
        "/brand-collections/brand-milgard-v250-style-line-series",
        "/window-type/milgard-vinyl-windows",
        "/brands/milgard",
    ],
    "andersen 100 series": [
        "/brand-collections/brand-andersen-100-series",
        "/window-type/andersen-vinyl-windows",
        "/brands/andersen",
    ],
    "marvin essential collection": [
        "/brand-collections/brand-marvin-essential-collection",
        "/window-type/marvin-fiberglass-windows",
        "/brands/marvin",
    ],
    "simonton daylightmax": [
        "/brand-collections/brand-simonton-daylightmax",
        "/window-type/simonton-vinyl-windows",
        "/brands/simonton",
    ],
    "milgard vs andersen windows": ["/brands/milgard", "/brands/andersen", "/brand"],
    "marvin vs andersen windows": ["/brands/marvin", "/brands/andersen", "/brand"],
    "simonton vs milgard windows": ["/brands/simonton", "/brands/milgard", "/brand"],
    # Brand seeds: overflow from /brands/* into brand+material and collections.
    "andersen windows": ["/brands/andersen", "/window-type/andersen-vinyl-windows", "/window-type/andersen-wood-windows"],
    "marvin windows": ["/brands/marvin", "/window-type/marvin-fiberglass-windows", "/window-type/wood-marvin-windows"],
    "milgard windows": ["/brands/milgard", "/window-type/milgard-vinyl-windows", "/window-type/milgard-fiberglass-windows"],
    "simonton windows": ["/brands/simonton", "/window-type/simonton-vinyl-windows"],
    "anlin windows": ["/brands/anlin", "/window-type/anlin-vinyl-windows"],
    "ply gem windows": ["/brands/ply-gem", "/window-type/ply-gem-vinyl-windows"],
    "alside windows": ["/brands/alside", "/window-type/alside-vinyl-windows"],
    "western window systems": ["/brands/western-window-systems", "/window-type/western-window-systems-aluminum-windows"],
    "andersen doors": ["/door-brands/andersen", "/door-types/andersen-fiberglass-doors", "/door-types/andersen-wood-doors"],
    "milgard doors": ["/door-brands/milgard", "/door-types/milgard-vinyl-doors", "/door-types/milgard-fiberglass-doors"],
    "marvin doors": ["/door-brands/marvin", "/door-types/marvin-wood-doors", "/door-types/marvin-fiberglass-doors"],
    "jeld wen doors": ["/door-brands/jeld-wen", "/door-types/jeld-wen-wood-doors", "/door-types/jeld-wen-vinyl-doors"],
    # Material seeds: overflow from /windows/* into brand+material pages.
    "vinyl windows": ["/windows/vinyl-windows", "/window-type/anlin-vinyl-windows", "/window-type/simonton-vinyl-windows"],
    "wood windows": ["/windows/wood-windows", "/window-type/andersen-wood-windows", "/window-type/jeld-wen-wood-windows"],
    "wood clad windows": ["/windows/wood-clad-windows", "/window-type/jeld-wen-wood-clad-windows"],
    "aluminum windows": ["/windows/aluminum-windows", "/window-type/milgard-aluminum-windows", "/window-type/western-window-systems-aluminum-windows"],
    "aluminum clad windows": ["/windows/aluminum-clad-windows", "/window-type/ply-gem-aluminum-clad-windows"],
    "fiberglass windows": ["/windows/fiberglass-windows", "/window-type/marvin-fiberglass-windows", "/window-type/milgard-fiberglass-windows"],
    "steel windows": ["/windows/steel-windows", "/window-type/italwindows-steel-windows"],
    # Local seeds: overflow into the county hub of the same city.
    "window replacement san francisco": ["/window-replacement/san-francisco", "/county-hub-pages/san-francisco-county"],
    "window replacement san jose": ["/window-replacement/san-jose", "/county-hub-pages/santa-clara-county"],
    "window replacement oakland": ["/window-replacement/oakland", "/county-hub-pages/alameda-county"],
    "window replacement fremont": ["/window-replacement/fremont", "/county-hub-pages/alameda-county"],
    "window replacement santa rosa": ["/window-replacement/santa-rosa", "/county-hub-pages/sonoma-county"],
    "window replacement walnut creek": ["/window-replacement/walnut-creek", "/county-hub-pages/contra-costa-county"],
}


if __name__ == "__main__":
    main()
